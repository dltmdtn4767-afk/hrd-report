"""
data_loader.py v2
엑셀 파일의 헤더를 자동 파싱하여 문항/카테고리를 인식하는 범용 로더
+ 전치(transposed) 구조 자동 감지
+ 다양한 헤더 패턴 지원 (L1, M1, 강사명 등)
"""
import re
import os
import openpyxl
from modules.config_manager import resolve_path


def _label_similarity(a: str, b: str) -> float:
    """두 문항 라벨의 유사도 (0~1). 간단한 bigram 기반."""
    if not a or not b:
        return 0.0
    a, b = a.lower().strip(), b.lower().strip()
    if a == b:
        return 1.0
    def bigrams(s):
        return set(s[i:i+2] for i in range(len(s)-1))
    ba, bb = bigrams(a), bigrams(b)
    if not ba or not bb:
        return 0.0
    return len(ba & bb) / max(len(ba | bb), 1)


def parse_course_info_from_filename(filepath):
    """파일명에서 고객사/과정명을 자동 추출"""
    basename = os.path.splitext(os.path.basename(filepath))[0]
    
    # 접두사 제거
    cleaned = re.sub(r'^[\[\(].*?[\]\)]\s*', '', basename).strip()
    cleaned = re.sub(r'\(\d{4}\)\s*$', '', cleaned).strip()
    cleaned = re.sub(r'[_\s]*엑스퍼트\s*(컨설팅)?\s*$', '', cleaned).strip()
    cleaned = re.sub(r'[_\s]*로우\s*데이터[_\s]*\d*\s*$', '', cleaned, flags=re.IGNORECASE).strip()
    
    # "회사명_과정명" 패턴
    m = re.match(r'^(.+?)[\s_]+(\d{4}년.+)$', cleaned)
    if m:
        return {"company": m.group(1).strip(), "course_name": m.group(2).strip(), "raw_name": cleaned}
    
    # "YYYY년 과정명" 패턴
    m = re.match(r'^(\d{4}년)\s+(.+?)\s+(.+)$', cleaned)
    if m:
        year = m.group(1)
        rest = m.group(2) + " " + m.group(3)
        words = rest.split()
        if len(words) >= 2:
            return {"company": words[0], "course_name": f"{year} {' '.join(words[1:])}", "raw_name": cleaned}
    
    # 회사명_과정명 (연도 없이)
    m = re.match(r'^(.+?)[_]+(.+)$', cleaned)
    if m:
        return {"company": m.group(1).strip(), "course_name": m.group(2).strip(), "raw_name": cleaned}
    
    return {"company": "", "course_name": cleaned, "raw_name": cleaned}


# ─── 주관식 키워드 (헤더 기반 1차 판별) ───
OPEN_ENDED_KEYWORDS = [
    '기술해', '작성해', '서술해', '입력', '명시해', '작성하', '기술하', '서술하',
    '의견', '개선', '제안', '감상', '담소', '담포', '희망', '바람', '요청',
    '자유롭게', '자유형식', '기타', '한가지를', '한 가지', '한가지',
    '목표를', '계획', '활용할', '적용할', '실천할', '노력', '이유를',
]

# 데이터 기반 재분류 임계값: None비율 50% 이상 시 정성 검토
OPEN_ENDED_NULL_THRESHOLD = 0.5

# 분석에서 제외할 레이블 블랙리스트 (데이터가 아닌 단순 타이틀/키워드)
LABEL_BLACKLIST = {
    "모듈", "교육 운영", "과정 전반", "평균", "합계", "기타", "교육 성과", 
    "교육 내용", "강사", "문항", "번호", "응답", "비중", "점수", "강사명"
}



def parse_header(header_text):
    """헤더 텍스트에서 문항 ID, 카테고리, 라벨을 자동 추출 (범용)"""
    header = str(header_text).strip()
    if not header:
        return None
    
    is_open_ended = any(kw in header for kw in OPEN_ENDED_KEYWORDS)
    
    # 패턴 1: "1-1. [교육 내용] 교육의 내용은..."
    m = re.match(r'^(\d+[-.]?\d*)[.\s]+\[(.+?)\]\s*(.+)$', header)
    if m:
        q_num = m.group(1).replace('.', '-')
        category = m.group(2).strip()
        label = m.group(3).strip()
        
        # [L1~L9], [M1~M9] 등 모듈 패턴
        if re.match(r'^[LMlm]\d+$', category):
            label = f"[{category}] {label}"
            category = "모듈"
        
        return {"id": f"Q{q_num}", "category": category, "label": label, "is_open_ended": is_open_ended}
    
    # 패턴 2: "5-1. 교육 후 내가 조직문화의..." 또는 "Q1-2. 문항내용"
    m = re.match(r'^([Qq]?\d+[-.]\d+)[.\s]+(.+)$', header)
    if m:
        id_str = m.group(1).upper()
        if not id_str.startswith('Q'): id_str = 'Q' + id_str
        id_str = id_str.replace('.', '-')
        
        label = m.group(2).strip()
        major = id_str.split('-')[0].replace('Q', '')
        category = f"분류{major}"
        return {"id": id_str, "category": category, "label": label, "is_open_ended": is_open_ended}
    
    # 패턴 3: "6. 만족한다" 또는 "Q6. 만족한다" (단일 번호)
    m = re.match(r'^([Qq]?\d+)[.\s]+(.+)$', header)
    if m:
        id_str = m.group(1).upper()
        if not id_str.startswith('Q'): id_str = 'Q' + id_str
        label = m.group(2).strip()
        if is_open_ended:
            return {"id": id_str, "category": "주관식", "label": label, "is_open_ended": True}
        else:
            return {"id": id_str, "category": "기타", "label": label, "is_open_ended": False}
    
    # 패턴 4: 번호 없는 텍스트 (길이 5자 이상이면 문항으로 간주)
    if len(header) >= 5 and not header.startswith('타임') and not header.startswith('응답'):
        # 숫자만도 아니고 날짜도 아닌 경우
        if not re.match(r'^\d+$', header) and not re.match(r'^\d{4}[-/]', header):
            return {"id": f"Q0", "category": "기타", "label": header, "is_open_ended": is_open_ended}
    
    return None


def _detect_orientation(ws):
    """
    엑셀 구조 감지: 일반(행=응답자) vs 전치(열=응답자)
    
    일반: 1행=헤더, 2행~=응답
    전치: 1열=문항ID, 2열~=응답자
    """
    # 1행의 값들 검사
    row1_vals = [ws.cell(row=1, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    row1_strs = [str(v).strip() if v else "" for v in row1_vals]
    
    # 1열의 값들 검사
    col1_vals = [ws.cell(row=r, column=1).value for r in range(1, min(15, ws.max_row + 1))]
    col1_strs = [str(v).strip() if v else "" for v in col1_vals]
    
    # 1행이 문항 헤더면 → 일반
    header_pattern_count = sum(1 for s in row1_strs if re.match(r'^([Qq]?\d+[-.])', s))
    if header_pattern_count >= 2: # 2개 이상만 있어도 데이터 시트로 간주
        return "normal"
        return "normal"
    
    # 1열이 문항처럼 보이면 → 일반 (문항이 세로)
    col1_header_count = sum(1 for s in col1_strs if re.match(r'^\d+[-.]', s))
    if col1_header_count >= 3:
        return "transposed"
    
    # 2행이 숫자 데이터면 → 일반
    row2_numeric = 0
    for c in range(2, min(15, ws.max_column + 1)):
        v = ws.cell(row=2, column=c).value
        try:
            float(v)
            row2_numeric += 1
        except:
            pass
    
    # 2열이 숫자 데이터면 → 전치
    col2_numeric = 0
    for r in range(2, min(15, ws.max_row + 1)):
        v = ws.cell(row=r, column=2).value
        try:
            float(v)
            col2_numeric += 1
        except:
            pass
    
    if row2_numeric > col2_numeric:
        return "normal"
    elif col2_numeric > row2_numeric:
        return "transposed"
    
    return "normal"


def _is_summary_row(row_values):
    """이 행이 결과 요약(평균, 합계 등) 행인지 판별"""
    summary_keywords = ['평균', '합계', '전체', '계', 'total', 'average', 'avg', 'sum', '결과']
    text_vals = [str(v).strip().lower() for v in row_values if v is not None]
    
    # 1. 키워드가 포함된 경우
    if any(any(kw in val for kw in summary_keywords) for val in text_vals):
        return True
    
    # 2. 정량 데이터열인데 소수점 값이 너무 많은 경우 (일반 응답은 1~5 정수)
    # (여기서는 개별 셀 수준이 아닌 행 전체의 특성으로 판단)
    return False

def _find_header_row(ws):
    """헤더가 있는 행 번호 찾기 (1~5행 검색)"""
    for row in range(1, min(6, ws.max_row + 1)):
        match_count = 0
        for col in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(row=row, column=col).value
            if v and parse_header(str(v)):
                match_count += 1
        if match_count >= 2:
            return row
    return 1


def load_from_excel(excel_path, sheet_name=None):
    """엑셀 파일에서 설문 응답 데이터를 범용으로 로드"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if sheet_name:
        try:
            ws = wb[sheet_name]
        except KeyError:
            ws = wb.active
    else:
        ws = wb.active
    
    actual_sheet = ws.title
    course_info = parse_course_info_from_filename(excel_path)
    
    # ─── 구조 감지 ───
    orientation = _detect_orientation(ws)
    
    if orientation == "transposed":
        result = _load_transposed(ws, course_info)
    else:
        result = _load_normal(ws, course_info)
    
    result["sheet_name"] = actual_sheet
    wb.close()
    return result


def _load_normal(ws, course_info):
    """일반 구조: 1행=헤더, 2행~=응답"""
    
    header_row = _find_header_row(ws)
    
    # 헤더 파싱
    headers = {}
    auto_id = 1
    empty_col_count = 0
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        
        if not val:
            empty_col_count += 1
            if empty_col_count >= 2: # 2행(줄) 이상의 빈 열이 나오면 중단
                break
            continue
            
        empty_col_count = 0
        header_text = str(val).strip()
        
        # 건너뛸 열
        skip_kw = ['타임스탬프', 'timestamp', '응답자', '이름', '이메일', 'email', '제출일']
        if any(kw in header_text.lower() for kw in skip_kw):
            continue
        
        parsed = parse_header(header_text)
        if parsed:
            if parsed["id"] == "Q0":
                parsed["id"] = f"Q{auto_id}"
            auto_id += 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            parsed["col"] = col_letter
            parsed["col_idx"] = col_idx
            parsed["header_raw"] = header_text
            headers[col_letter] = parsed
    
    if not headers:
        raise ValueError(f"유효한 문항 헤더를 찾을 수 없습니다. (1행: {[ws.cell(row=header_row, column=c).value for c in range(1, min(5, ws.max_column+1))]})")
    
    # 응답 데이터 수집
    data_start = header_row + 1
    
    for col_letter, q in headers.items():
        q["scores"] = []
        q["answers"] = []
    
    row_idx = data_start
    empty_count = 0
    while row_idx <= ws.max_row and empty_count < 2: # 2줄 이상의 떨어진 데이터는 무시
        # 이 행의 모든 값을 가져와 요약 행인지 확인
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if _is_summary_row(row_values):
            row_idx += 1
            continue

        # 아무 문항 열에서든 데이터가 있는지 확인
        has_data = False
        for col_letter, q in headers.items():
            val = ws.cell(row=row_idx, column=q["col_idx"]).value
            if val is not None:
                txt = str(val).strip()
                if txt == "" or txt in LABEL_BLACKLIST:
                    continue
                has_data = True
                break
        
        if not has_data:
            empty_count += 1
            row_idx += 1
            continue
        
        for col_letter, q in headers.items():
            val = ws.cell(row=row_idx, column=q["col_idx"]).value
            txt_val = str(val).strip() if val is not None else ""
            
            # 블랙리스트 필터링
            if txt_val in LABEL_BLACKLIST:
                continue

            if q["is_open_ended"]:
                if val and txt_val:
                    q["answers"].append(txt_val)
            else:
                try:
                    num_val = float(val)
                    if num_val != int(num_val) and not q.get("is_floating_allowed", False):
                        pass
                    q["scores"].append(num_val)
                except (TypeError, ValueError):
                    q["scores"].append(None)
                    raw = txt_val
                    if raw and raw not in ("-", "0", "n/a", "N/A", "none"):
                        q.setdefault("raw_texts", []).append(raw)
        
        row_idx += 1
    
    response_count = sum(1 for q in headers.values() if q["scores"] or q["answers"]) # 실제 데이터가 있는 행 기준
    if response_count <= 0:
        raise ValueError("설문 응답 데이터가 없습니다.")
    
    return _build_result(headers, response_count, course_info)


def _load_transposed(ws, course_info):
    """전치 구조: 행=문항, 열=응답자"""
    
    headers = {}
    auto_id = 1
    
    # 문항 헤더 찾기 (1열 또는 2열에서)
    question_col = 1
    for r in range(1, min(30, ws.max_row + 1)):
        val = ws.cell(row=r, column=1).value
        if val and parse_header(str(val)):
            question_col = 1
            break
    
    # 데이터 시작 열 찾기
    data_start_col = 2
    for c in range(2, min(10, ws.max_column + 1)):
        v = ws.cell(row=4, column=c).value  # 4행쯤에서 숫자 확인
        try:
            float(v)
            data_start_col = c
            break
        except:
            continue
    
    # 각 행을 문항으로
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=question_col).value
        if not val:
            continue
        header_text = str(val).strip()
        
        skip_kw = ['응답자', '시작일', '종료일', '제출일', 'id']
        if any(kw in header_text.lower() for kw in skip_kw):
            continue
        
        parsed = parse_header(header_text)
        if not parsed:
            continue
        
        if parsed["id"] == "Q0":
            parsed["id"] = f"Q{auto_id}"
        auto_id += 1
        
        # 이 행의 데이터 (열 방향)
        parsed["scores"] = []
        parsed["answers"] = []
        
        for c in range(data_start_col, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if parsed["is_open_ended"]:
                if val and str(val).strip():
                    parsed["answers"].append(str(val).strip())
            else:
                try:
                    parsed["scores"].append(float(val))
                except (TypeError, ValueError):
                    parsed["scores"].append(None)
                    raw = str(val).strip() if val is not None else ""
                    if raw and raw not in ("-", "0", "n/a", "N/A", "none"):
                        parsed.setdefault("raw_texts", []).append(raw)
        
        parsed["col"] = str(r)
        parsed["col_idx"] = r
        parsed["header_raw"] = header_text
        headers[str(r)] = parsed
    
    if not headers:
        raise ValueError("전치 구조에서 유효한 문항을 찾을 수 없습니다.")
    
    response_count = ws.max_column - data_start_col + 1
    return _build_result(headers, response_count, course_info)


def _build_result(headers, response_count, course_info):
    """공통 결과 빌드"""
    questions = []
    open_ended = []
    instructor_names = set()
    
    for key, q in headers.items():
        # ─── 데이터 기반 사후 재분류 (정밀 판별) ───
        if not q["is_open_ended"]:
            scores = q.get("scores", [])
            raw_texts = q.get("raw_texts", [])
            valid_scores = [s for s in scores if s is not None]
            
            total = len(scores)
            score_count = len(valid_scores)
            text_count = len(raw_texts)
            null_count = scores.count(None)
            null_ratio = null_count / total if total > 0 else 1.0
            
            # 의미 있는 텍스트 추출 (블랙리스트 제외)
            meaningful_texts = [t for t in raw_texts if len(str(t).strip()) >= 2 and t not in LABEL_BLACKLIST]
            
            # [수치 데이터 우선 판별 원칙 강화]
            # 1. 1~5점 사이의 점수가 1개라도 있으면 일단 정량 후보
            # 2. 정량 데이터가 전체 유효 데이터(수치+텍스트)의 30% 이상이면 정량으로 유지
            has_valid_range_score = any(1 <= s <= 5 for s in valid_scores)
            total_valid_count = score_count + text_count
            
            # 수치 데이터가 어느 정도 존재하거나, 1-5점 패턴이 명확한 경우
            is_likely_quant = has_valid_range_score or (score_count > 0 and (score_count / max(total_valid_count, 1) >= 0.3))

            if not is_likely_quant:
                # 수치가 거의 없거나 1-5 범위를 벗어나는 경우에만 정성으로 전환 검토
                if len(meaningful_texts) > 0 or (null_ratio >= OPEN_ENDED_NULL_THRESHOLD and text_count > 0):
                    q["is_open_ended"] = True
                    q["answers"] = raw_texts
            else:
                # 정량으로 판별된 경우, 발견된 텍스트들은 'metadata' 성격의 응답으로 저장 (차후 프리뷰 활용 가능)
                if meaningful_texts:
                    q["comment_metadata"] = meaningful_texts

        if q["is_open_ended"]:
            # 정성 응답에서도 블랙리스트 제거
            clean_answers = [a for a in q.get("answers", []) if a not in LABEL_BLACKLIST]
            open_ended.append({"id": q["id"], "label": q["label"], "answers": clean_answers})
        else:
            valid = [s for s in q.get("scores", []) if s is not None]
            avg = round(sum(valid) / len(valid), 2) if valid else 0
            questions.append({
                "col": q.get("col", ""), "id": q["id"], "category": q["category"],
                "label": q["label"], "avg": avg, "count": len(valid),
            })
            if "강사" in q["category"]:
                instructor_names.add(q["category"])
    
    # 대분류별 평균
    categories_order = []
    for q in questions:
        if q["category"] not in categories_order:
            categories_order.append(q["category"])
    
    categories = []
    for cat_name in categories_order:
        cat_qs = [q for q in questions if q["category"] == cat_name]
        cat_avg = round(sum(q["avg"] for q in cat_qs) / len(cat_qs), 2) if cat_qs else 0
        categories.append({"name": cat_name, "questions": cat_qs, "avg": cat_avg})
    
    all_avgs = [q["avg"] for q in questions if q["avg"] > 0]
    overall_average = round(sum(all_avgs) / len(all_avgs), 2) if all_avgs else 0
    
    chart_labels = []
    for q in questions:
        lbl = q["label"]
        if len(lbl) > 15:
            lbl = lbl[:12] + "..."
        chart_labels.append(lbl)
    
    return {
        "course_info": course_info,
        "response_count": response_count,
        "questions": questions,
        "categories": categories,
        "categories_order": categories_order,
        "overall_average": overall_average,
        "open_ended": open_ended,
        "chart_data": {"labels": chart_labels, "values": [q["avg"] for q in questions]},
        "instructor_names": list(instructor_names),
    }


def get_available_sheets(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def get_summary_text(data):
    ci = data["course_info"]
    lines = [
        f"=== {ci['company']} {ci['course_name']} 만족도 분석 ({data.get('sheet_name', '')}) ===",
        f"",
        f"* 응답 인원: {data['response_count']}명",
        f"* 문항 수: {len(data['questions'])}개 (주관식 {len(data['open_ended'])}개)",
        f"* 대분류: {len(data['categories'])}개",
        f"* 전체 평균: {data['overall_average']}점 / 5점",
        f"",
        f"--- 대분류별 평균 ---",
    ]
    for cat in data["categories"]:
        lines.append(f"  - {cat['name']}: {cat['avg']}점 ({len(cat['questions'])}문항)")
    
    if data["open_ended"]:
        lines.append("")
        lines.append("--- 주관식 ---")
        for oe in data["open_ended"]:
            lines.append(f"  [{oe['id']}] ({len(oe['answers'])}건)")
    
    return "\n".join(lines)


# ════════════════════════════════════════════════════
# 멀티 시트 로딩 (차수별 분리 + 종합)
# ════════════════════════════════════════════════════

def _is_data_sheet(ws):
    """이 시트가 설문 데이터를 포함하는지 확인"""
    # 최소 2행, 2열 이상
    if ws.max_row < 2 or ws.max_column < 2:
        return False
    # 1행에 인식 가능한 헤더가 1개 이상 (최소 조건 완화)
    header_count = 0
    for col in range(1, min(30, ws.max_column + 1)):
        v = ws.cell(row=1, column=col).value
        if v and parse_header(str(v)):
            header_count += 1
    return header_count >= 1


def load_all_sheets(excel_path):
    """
    모든 시트를 로드하여 차수별 데이터 + 종합 데이터 반환
    
    Returns:
        {
          "course_info": {...},
          "sessions": [          # 차수별
              {"sheet_name": "1차수", "data": {...}},
              ...
          ],
          "combined": {...},     # 종합 (모든 차수 합산)
          "multi_session": True,
        }
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    
    course_info = parse_course_info_from_filename(excel_path)
    
    sessions = []
    for sheet_name in sheets:
        try:
            wb2 = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb2[sheet_name]
            
            if not _is_data_sheet(ws):
                wb2.close()
                continue
            
            orientation = _detect_orientation(ws)
            if orientation == "transposed":
                result = _load_transposed(ws, course_info)
            else:
                result = _load_normal(ws, course_info)
            
            result["sheet_name"] = sheet_name
            wb2.close()
            
            # 차수 번호 추출 (1차수, 2차수, 1회차, etc.)
            session_num = _extract_session_number(sheet_name)
            result["session_label"] = session_num or sheet_name
            
            sessions.append(result)
        except Exception as e:
            print(f"시트 '{sheet_name}' 로딩 실패: {e}")
            continue
    
    if not sessions:
        # 폴백: 첫 번째 시트만
        single = load_from_excel(excel_path)
        return {
            "course_info": course_info,
            "sessions": [single],
            "combined": single,
            "multi_session": False,
        }
    
    combined = _combine_sessions(sessions, course_info)
    
    return {
        "course_info": course_info,
        "sessions": sessions,
        "combined": combined,
        "multi_session": len(sessions) > 1,
    }


def _extract_session_number(sheet_name):
    """시트 이름에서 차수/회차 번호 추출"""
    # "1차수", "2차수", "1회차", "1기", "1차" 등
    m = re.search(r'(\d+)\s*(차수|회차|기|차|round)', sheet_name, re.IGNORECASE)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    m = re.search(r'(\d+)', sheet_name)
    if m:
        return f"{m.group(1)}차수"
    return sheet_name


def _combine_sessions(sessions, course_info):
    """여러 차수의 데이터를 종합 집계"""
    # 공통 문항 ID 기준으로 병합
    all_q_ids = []
    for s in sessions:
        for q in s["questions"]:
            if q["id"] not in all_q_ids:
                all_q_ids.append(q["id"])
    
    combined_questions = []
    for qid in all_q_ids:
        # 각 세션에서 해당 문항 수집
        per_session = []
        for s in sessions:
            q = next((qq for qq in s["questions"] if qq["id"] == qid), None)
            if q:
                per_session.append(q)
        
        if not per_session:
            continue
        
        # 평균의 평균 (세션별 응답 수 가중)
        total_count = sum(q["count"] for q in per_session)
        weighted_avg = sum(q["avg"] * q["count"] for q in per_session) / total_count if total_count > 0 else 0
        
        base = per_session[0].copy()
        base["avg"] = round(weighted_avg, 2)
        base["count"] = total_count
        base["per_session"] = {s["session_label"]: q["avg"] for s, q in zip(sessions, per_session) if q}
        combined_questions.append(base)
    
    # 카테고리 재구성
    categories_order = []
    for q in combined_questions:
        if q["category"] not in categories_order:
            categories_order.append(q["category"])
    
    categories = []
    for cat_name in categories_order:
        cat_qs = [q for q in combined_questions if q["category"] == cat_name]
        cat_avg = round(sum(q["avg"] for q in cat_qs) / len(cat_qs), 2) if cat_qs else 0
        # 카테고리별 세션별 평균도
        cat_per_session = {}
        for s in sessions:
            s_qs = [q for q in s["questions"] if q["category"] == cat_name]
            if s_qs:
                cat_per_session[s["session_label"]] = round(sum(q["avg"] for q in s_qs)/len(s_qs), 2)
        categories.append({
            "name": cat_name, "questions": cat_qs, "avg": cat_avg,
            "per_session": cat_per_session,
        })
    
    # 주관식 병합 — 3중 검증 (id + label 유사도 + 응답 내용 일관성)
    all_oe_ids = []
    combined_oe = []
    for s in sessions:
        for oe in s.get("open_ended", []):
            oe_id = oe["id"]
            oe_label = oe["label"].strip()
            oe_answers = list(oe.get("answers", []))

            # ── 검증 1: id 기반 매칭 ──
            existing = None
            for x in combined_oe:
                if x["id"] == oe_id:
                    existing = x
                    break

            if existing:
                # ── 검증 2: label 유사도 확인 ──
                ex_label = existing["label"].strip()
                label_match = (
                    oe_label == ex_label or
                    oe_label in ex_label or
                    ex_label in oe_label or
                    _label_similarity(oe_label, ex_label) >= 0.6
                )
                if label_match:
                    # ── 검증 3: 중복 응답 제거 후 병합 ──
                    existing_set = set(existing["answers"])
                    new_answers = [a for a in oe_answers if a not in existing_set]
                    existing["answers"].extend(new_answers)
                    print(f"[QA-MERGE] {oe_id} '{oe_label}' 병합 (+{len(new_answers)}건)")
                else:
                    # label 불일치 → 별도 문항으로 분리
                    new_id = f"{oe_id}_{s.get('session_label','')}"
                    combined_oe.append({"id": new_id, "label": oe_label, "answers": oe_answers})
                    print(f"[QA-SPLIT] {oe_id} label 불일치: '{ex_label}' vs '{oe_label}' → {new_id}")
            else:
                combined_oe.append({"id": oe_id, "label": oe_label, "answers": oe_answers})
    
    total_resp = sum(s["response_count"] for s in sessions)
    all_avgs = [q["avg"] for q in combined_questions if q["avg"] > 0]
    overall_avg = round(sum(all_avgs) / len(all_avgs), 2) if all_avgs else 0
    
    chart_labels = [q["label"][:12] + "..." if len(q["label"]) > 15 else q["label"] for q in combined_questions]
    
    # 세션별 종합 차트 데이터
    session_chart = {
        "labels": [s["session_label"] for s in sessions] + ["종합"],
        "datasets": [],
    }
    for cat in categories:
        dataset = {
            "label": cat["name"],
            "values": [cat["per_session"].get(s["session_label"], 0) for s in sessions] + [cat["avg"]],
        }
        session_chart["datasets"].append(dataset)
    
    return {
        "course_info": course_info,
        "sheet_name": "종합",
        "session_label": "종합",
        "response_count": total_resp,
        "questions": combined_questions,
        "categories": categories,
        "categories_order": categories_order,
        "overall_average": overall_avg,
        "open_ended": combined_oe,
        "chart_data": {"labels": chart_labels, "values": [q["avg"] for q in combined_questions]},
        "instructor_names": list(set(n for s in sessions for n in s.get("instructor_names", []))),
        "session_chart": session_chart,
        "session_count": len(sessions),
    }

