"""업로드된 엑셀 헤더 분석"""
import sys, glob, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

for d in sorted(glob.glob('uploads/*/'))[:5]:
    files = glob.glob(os.path.join(d, '*.xlsx'))
    if not files: continue
    f = files[0]
    bn = os.path.basename(f)
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        print(f"\n=== {bn[:50]} ===")
        print(f"  시트: {ws.title}, 행: {ws.max_row}, 열: {ws.max_column}")
        # 1-3행 출력
        for row in range(1, min(4, ws.max_row+1)):
            vals = []
            for col in range(1, min(8, ws.max_column+1)):
                v = ws.cell(row=row, column=col).value
                vals.append(str(v)[:25] if v else "")
            print(f"  행{row}: {vals}")
        wb.close()
    except Exception as e:
        print(f"\n=== {bn[:50]} === 에러: {e}")
